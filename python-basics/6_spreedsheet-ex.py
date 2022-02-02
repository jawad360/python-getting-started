import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
company_value = {}
product_10 = {}

# range(number) is an iterator starting from 0 to number - 1
# range(index, number) will iterate starting from index to number - 1
for product_row in range(2, product_list.max_row + 1):
    # product_row will contain the index of the current range
    # cell will take actual row number and not the index

    # Exercise 1
    supplier_name = product_list.cell(product_row, 4).value
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

    # Exercise 2
    price = product_list.cell(product_row, 3).value
    inv_count = product_list.cell(product_row, 2).value
    if supplier_name in company_value:
        company_value[supplier_name] += price * inv_count
    else:
        company_value[supplier_name] = price * inv_count

    # Exercise 3
    product = product_list.cell(product_row, 1).value
    if inv_count < 10:
        product_10[product] = inv_count

    # Exercise 4
    inv_price = product_list.cell(product_row, 5)
    inv_price.value = price * inv_count

print(product_per_supplier)
print(company_value)
print(product_10)

inv_file.save("inventory_total.xlsx")



